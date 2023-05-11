import xlsxwriter
from openpyxl import load_workbook
import os.path as path
import graphal as ga
import time


def make_test(size=None,
              filename='test.xlsx',
              importfilename=None,
              exporttoexcel=True,
              exporttofile=False,
              printmatrix=False,
              printcolour=False,
              printamount=False,
              printtime=False,
              iterations=1):
    if size is None or size[1] < size[0]-1 or size[1] > size[0]*(size[0]-1)/2:
        exit("Size is bad")
    if not path.exists(filename) and exporttoexcel:
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet('Sheet1')
        workbook.close()
        wb = load_workbook(filename=filename)
        ws = wb['Sheet1']
        ws["C1"] = "Optymalny"
        ws["E1"] = "Algorytm LF"
        ws["A2"] = "Wierzchołki"
        ws["B2"] = "Krawędzie"
        ws["C2"] = "Czas"
        ws["D2"] = "Liczba kolorów"
        ws["E2"] = "Czas"
        ws["F2"] = "Liczba kolorów"
        wb.save(filename)
    if importfilename is not None:
        iterations = 1
    for _ in range(0, iterations):
        if importfilename is None:
            matrix = ga.make(size[0], size[1])
        else:
            matrix = ga.import_matrix(importfilename)
        if matrix is None:
            exit("Creating matrix is fault. Bad importfilename")
        start = time.perf_counter()
        colourlf = ga.coloring(matrix)
        end = time.perf_counter()
        time_lf = end - start
        start = time.perf_counter()
        colouro = ga.optimal(matrix)
        end = time.perf_counter()
        time_optimal = end - start
        if exporttoexcel:
            wb = load_workbook(filename=filename)
            ws = wb['Sheet1']
            row = 3
            while not ws["A" + str(row)].value is None:
                row += 1
            ws["A" + str(row)] = str(size[0])
            ws["B" + str(row)] = str(size[1])
            ws["C" + str(row)] = time_optimal
            ws["D" + str(row)] = max(colouro)+1
            ws["E" + str(row)] = time_lf
            ws["F" + str(row)] = max(colourlf)+1
            wb.save(filename)
            print("Test exported successful")
        if exporttofile and importfilename is None:
            ga.export_matrix(matrix)
            print("Graph exported successful")
        if printmatrix:
            print("Matrix of incidents:")
            print(matrix)
        if printcolour:
            print("Optimal algorithm: ", colouro)
            print("Algorithm LF:", colourlf)
        if printamount:
            print("Optimal algorithm: ", max(colouro)+1)
            print("Algorithm LF:", max(colourlf)+1)
        if printtime:
            print("Optimal algorithm: ", time_optimal, "s")
            print("Algorithm LF:", time_lf, "s")
            print("Delta time:", time_optimal-time_lf, "s")
        print("Done")